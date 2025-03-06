import openpyxl
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from os.path import join, abspath

# Calculating the Chézy roughness coefficient C value
# using an ensemble of trained ANNs (ANN_A, ANN_B1, ANN_B2),
# weight matrices of such neural networks are loaded from text files,
# input arguments are loaded from a MS Excel file,
# the calculation results are stored in a MS Excel file

# defining a function to calculate the derivative of the activation function

def logistic(x):
    return 1.0 / (1 + np.exp(-x))

# defining a function to calculate ANN outputs for ANN_A, ANN_B1, ANN_B2

def calculating_ANN(ANN_ID):
    # calculation of approximate values of the Chezy coefficient C_A, or the values of its refinement  delta1, delta2
    # respectively using trained neural networks ANN_A, ANN_B1, ANN_B2,

    # reading of matrix of weight coefficients W_1 і W_2

    data_path = join('.', 'Data', "weights_matrix_1_"+ANN_ID+".txt")
    data_path = abspath(data_path)

    f = open(data_path)
    raw_matrix1 = f.readlines()
    f.close()

    data_path = join('.', 'Data', "weights_matrix_2_"+ANN_ID+".txt")
    data_path = abspath(data_path)

    f = open(data_path)
    raw_matrix2 = f.readlines()
    f.close()

    # initialization of the neural network output vector
    output_ANN = np.zeros(shape=(max_row_data))

    # initialization of matrices W_1 і W_2
    W_1 = np.zeros(
        shape=(input_size, hidden_size))
    W_2 = np.zeros(shape=(hidden_size, output_size))

    # filling of matrices W_1 і W_2

    for i in range(len(raw_matrix1)):
        b = raw_matrix1[i]
        a = b.split(' ')
        for j in range(len(raw_matrix2)):
            W_1[i][j] = float(a[j])

    for i in range(len(raw_matrix2)):
        W_2[i] = float(raw_matrix2[i])

    if (len(raw_matrix1) == input_size) and (len(raw_matrix2) == hidden_size):
        param_matching = 0
        print('The input data for ANN_' + ANN_ID + ' correspond to the parameters of the ANN matrices of weight,')
    else:
        param_matching = 1
        print('The input data do not correspond to the ANN matrices of weights.')
        print('Check that the number of inputs and neurons in the hidden layer match')
        print(' to the parameters of the trained network in the input file:')
        print('the number of columns W_1 (the number of neurons in the hidden layer of the network) =', len(raw_matrix2))
        print('the number of rows W_1 (the number of the ANN inputs) =', len(raw_matrix1))
        print('the number of rows W_2 (the number of neurons in the hidden layer) =', len(raw_matrix2))

    # calculation of the network outputs
    if param_matching == 0:
        for i in range(max_row_data):
            # direct course of calculations
            layer_0 = characteristics_riverbed[i:i + 1]
            layer_1 = logistic(np.dot(layer_0, W_1))
            layer_2 = np.dot(layer_1, W_2)
            output_ANN[i] = layer_2[0][0]
            print('for a set of parameters №', i + 1, ', calculated output for ANN_'+ANN_ID+ ' = ', output_ANN[i])

    return  output_ANN

# ***************** end of function calculating_ANN *****************

# defining a function for combining predictions from an ensemble of neural networks ANN_A, ANN_B1, ANN_B2
def aggregation_C(C, C1, C2, Q, B, H, Sf):
    # Prediction aggregation is based on the max voting method,
    # analysis of predictions (C_A, C_B1, C_B2) and identifying the best of the predictions (inverse problem),
    # in this case the value C_A, C_B1, C_B2 needs to be converted, because they were obtained using a neural network,
    # so we multiply them by 100 when calling the function, i.e. C = C_A*100, C1 = C_B1*100, C2 = C_B2*100

    delta_V = np.zeros(shape=(3))
    V_Q = Q / (B * H) # reference value of water flow velocity V(Q)
    V = C * (H * Sf) ** 0.5 # the value of the flow velocity V(C_A), calculated by the Chezy formula taking into account C_A
    V1 = C1 * (H * Sf) ** 0.5 # value of water flow velocity calculated taking into account C_B1
    V2 = C2 * (H * Sf) ** 0.5 # value of water flow velocity calculated taking into account C_B2

    # calculation of the deviation from the reference value of the water flow velocity V(Q)
    delta_V[0] = abs(V_Q - V)
    delta_V[1] = abs(V_Q - V1)
    delta_V[2] = abs(V_Q - V2)

    # minimal deviation detection
    min_delta = delta_V[0]
    min_ind = 0
    for i in range(1,3):
        if min_delta > delta_V[i]:
            min_delta = delta_V[i]
            min_ind = i

    # the output of the neural network with the minimum deviation from the reference value is selected
    if min_ind == 0: return C/100 # output = C_A
    if min_ind == 1: return C1/100 # output = C_B1
    if min_ind == 2: return C2/100 # output = C_B2

# ***************** end of function aggregation_C *****************

# definition and initialization of ANN parameters and input data vectors
# open the specified file only to read the values
data_path = join('.',"Input.xlsx")
data_path = abspath(data_path) # it is created a full path to open the file

# the directory in which the program is running
Input_Data = openpyxl.open(data_path, read_only=True, data_only=True)
sheet_data = Input_Data.worksheets[0]

max_row_data = sheet_data['A3'].value  # number of input data examples

# network parameters corresponding to the input data
input_size = sheet_data['A5'].value  # number of inputs
hidden_size = sheet_data['A7'].value  # number of neurons in the hidden layer
output_size = sheet_data['A9'].value  # number of outputs

# definition and initialization of matrices of the ANN input values and
# matrix of the Chézy roughness coefficient calculated values
# and initialization of vectors for non-normalized parameters Q, B, H, Sf

len_ryadok = input_size

characteristics_riverbed = np.zeros(
    shape=(max_row_data, len_ryadok))  # input - normalized hydrodynamic characteristics of the channel

coef_C = np.zeros(shape=(max_row_data))  # output - the value of the Chézy roughness coefficient, established using an ensemble of neural networks
coef_C_A = np.zeros(shape=(max_row_data))  # intermediate output - the value of the Chézy roughness coefficient calculated using ANN_A
coef_C_B1 = np.zeros(shape=(max_row_data))  # intermediate output - the value of the Chézy roughness coefficient calculated using ANN_B1
coef_C_B2 = np.zeros(shape=(max_row_data))  # intermediate output - the value of the Chézy roughness coefficient calculated using ANN_B2
coef_delta1 = np.zeros(shape=(max_row_data))  #intermediate output - refinement for the Chézy roughness coefficient, the result of the ANN_B1 network calculations
coef_delta2 = np.zeros(shape=(max_row_data))  # intermediate output - refinement for the Chézy roughness coefficient, the result of the ANN_B2 network calculations

# parameters required to solve the inverse problem of calculating the average flow velocity V(Q), V(C_A), V(C_B1), V(C_B2)
Q = np.zeros(shape=(max_row_data))  # the water discharge
B = np.zeros(shape=(max_row_data)) # the average flow width
H = np.zeros(shape=(max_row_data)) # the average flow depth
Sf = np.zeros(shape=(max_row_data)) # the water surface slope

param_matching = 0 # error tracking (0 - all is well, 1 - error)
# param_matching checks whether the parameters of the neural network correspond to the parameters of the loaded weight coefficient matrices

# loading input values
for i in range(max_row_data):
    Q[i] = sheet_data[i + 2][input_size + 2].value
    B[i] = sheet_data[i + 2][input_size + 3].value
    H[i] = sheet_data[i + 2][input_size + 4].value
    Sf[i] = sheet_data[i + 2][input_size + 5].value

    for j in range(len_ryadok):
        characteristics_riverbed[i][j] = sheet_data[i + 2][j + 1].value

print('The input data for the ANN was downloaded from the file:')
print(data_path)
print('Parameters of neural networks ANN_A, ANN_B1, ANN_B2:')
print('number of inputsв =', input_size)
print('number of neurons in the hidden layer =', hidden_size)
print('number of outputs =', output_size)
print('Number of input examples:', max_row_data)

list_param=[] # list of descriptions of input parameters
print('The C/100 coefficient is investigated taking into account the parameters', end=' ')
for j in range(len_ryadok):
    list_param.append(sheet_data[1][j+1].value)
    print(sheet_data[1][j+1].value, end=',')
print()

Input_Data.close() # closing of the MS Excel input data file

#***************

# calculating predictions of trained neural networks ANN_A, ANN_B1, ANN_B2
coef_C_A = calculating_ANN("A")  # approximate prediction of neural network ANN_A
coef_delta1 = calculating_ANN("B1")  # approximate prediction of neural networkю ANN_B1
coef_delta2 = calculating_ANN("B2")   # approximate prediction of neural network ANN_B2

# calculating the outputs of trained neural networks  ANN_B1 та ANN_B2
print('Calculation of the values of the Chézy roughness coefficient taking into account the predictions of the neural networks ANN_B1 and ANN_B2 (normalized values):')
for i in range(max_row_data):
    coef_C_B1[i] = coef_C_A[i] + coef_delta1[i]
    coef_C_B2[i] = coef_C_A[i] - coef_delta2[i]
    print('coef_C_A[', i, '] = ', coef_C_A[i])
    print('coef_C_B1[', i, '] = ', coef_C_B1[i])
    print('coef_C_B2[', i, '] = ', coef_C_B2[i])

# Ensemble combination of predictions of neural networks ANN_A, ANN_B1, ANN_B2 (the inverse problem is considered),
print('Aggregation of predictions of neural networks ANN_A, ANN_B1, ANN_B2 into a single output:')
for i in range(max_row_data):
    # Prediction aggregation
    coef_C[i] = aggregation_C(coef_C_A[i]*100, coef_C_B1[i]*100, coef_C_B2[i]*100, Q[i], B[i], H[i], Sf[i])
    print('coef_C[', i, '] = ', coef_C[i])

# ***************

if param_matching == 0:
    # creating a MS Excel file to write data to a spreadsheet
    Out_Data = Workbook()
    sheet_data = Out_Data.active
    sheet_data.title = "output_data"

    # recording data in the table
    # column names are added
    sheet_data['A1'].value = 'j'
    sheet_data.cell(row=1, column=len_ryadok + 2).value = 'C/100'
    for j in range(len_ryadok):
        sheet_data[1][j + 1].value = list_param[j]

    # recording the input data and the corresponding calculated values of C
    for i in range(max_row_data):
        sheet_data['A' + str(i + 2)].value = i + 1
        sheet_data.cell(row=i + 2, column=len_ryadok + 2).value = coef_C[i]
        for j in range(len_ryadok):
            sheet_data[i + 2][j + 1].value = characteristics_riverbed[i][j]

    # formatting the output to the table
    for i in range(1, len_ryadok + 2):
        zag = sheet_data.cell(row=1, column=i)
        zag.alignment = Alignment(horizontal='center')
        zag.font = Font(bold=True, italic=False, color='DC143C', size=12)

        zag = sheet_data.cell(row=1, column=len_ryadok + 2)
        zag.alignment = Alignment(horizontal='center')
        zag.font = Font(bold=True, italic=False, color='000000', size=12)

    for i in range(max_row_data + 1):
        sheet_data.cell(row=i + 1, column=len_ryadok + 2).fill = PatternFill(
            fill_type='solid', start_color='90EE90', end_color='90EE90'
        )

    exfilename = join('.', ('Output.xlsx'))  # creating (join from parts) file name for storage
    exfilename = abspath(exfilename)  # creating a complete path to store the file

    Out_Data.save(exfilename)  # storing the file
    Out_Data.close()
    print("Saving the calculation results in the file: ")
    print(exfilename)

input('Press any key to end the program.')